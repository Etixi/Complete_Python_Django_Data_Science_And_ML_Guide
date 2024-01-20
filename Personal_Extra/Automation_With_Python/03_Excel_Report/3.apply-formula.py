from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('data/barchart.xlsx')
sheet = wb['Sheet1']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Write an Excel formula with Python
# sheet['B8'] = '=SUM(B6:B7)'
# sheet['B8'].style = 'Currency'

# Write multiple formulas with a for loop
for i in range(min_column+1, max_column+1):  # (B, G+1)
    print(i)
    letter = get_column_letter(i)
    # print(letter)
    # sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    print(f'=SUM({letter}{min_row + 1}:{letter}{max_row})')
    sheet[f'{letter}{max_row + 1}'].style = 'Currency'


# wb.save('data/report.xlsx')